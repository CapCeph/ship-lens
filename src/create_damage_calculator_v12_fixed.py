#!/usr/bin/env python3
"""
Creates a Star Citizen Damage Calculator Excel workbook v12.
FIXED VERSION with proper ship naming and AI variant filtering.
"""

import csv
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DATA_ROOT = Path.home() / "Projects" / "sc-data-verification" / "processed"
OUTPUT_FILE = Path.home() / "Projects" / "sc-data-verification" / "Star_Citizen_Damage_Calculator_v12.xlsx"

MAX_WEAPON_SLOTS = 10

# Manufacturer mappings
MANUFACTURERS = {
    'AEGS': 'Aegis', 'ANVL': 'Anvil', 'ARGO': 'Argo', 'BANU': 'Banu',
    'BEHR': 'Behring', 'CNOU': 'Consolidated Outland', 'CRUS': 'Crusader', 'DRAK': 'Drake',
    'ESPR': 'Esperia', 'GAMA': 'Gatac', 'GLSN': 'Gallenson', 'KRIG': 'Kruger',
    'MISC': 'MISC', 'MRAI': 'Mirai', 'ORIG': 'Origin', 'RSI': 'RSI',
    'TMBL': 'Tumbril', 'VNCL': 'Vanduul', 'XIAN': "Xi'An", 'XNAA': "Xi'An",
}

# Ship model name corrections for proper capitalization
SHIP_NAME_FIXES = {
    'avenger': 'Avenger',
    'stalker': 'Stalker',
    'titan': 'Titan',
    'warlock': 'Warlock',
    'renegade': 'Renegade',
    'eclipse': 'Eclipse',
    'gladius': 'Gladius',
    'valiant': 'Valiant',
    'hammerhead': 'Hammerhead',
    'idris': 'Idris',
    'javelin': 'Javelin',
    'reclaimer': 'Reclaimer',
    'redeemer': 'Redeemer',
    'retaliator': 'Retaliator',
    'sabre': 'Sabre',
    'comet': 'Comet',
    'firebird': 'Firebird',
    'peregrine': 'Peregrine',
    'raven': 'Raven',
    'vanguard': 'Vanguard',
    'harbinger': 'Harbinger',
    'hoplite': 'Hoplite',
    'sentinel': 'Sentinel',
    'arrow': 'Arrow',
    'asgard': 'Asgard',
    'pisces': 'Pisces',
    'expedition': 'Expedition',
    'carrack': 'Carrack',
    'gladiator': 'Gladiator',
    'hawk': 'Hawk',
    'hornet': 'Hornet',
    'hurricane': 'Hurricane',
    'lightning': 'Lightning',
    'paladin': 'Paladin',
    'terrapin': 'Terrapin',
    'medic': 'Medic',
    'valkyrie': 'Valkyrie',
    'mole': 'MOLE',
    'carbon': 'Carbon',
    'talus': 'Talus',
    'mpuv': 'MPUV',
    'transport': 'Transport',
    'raft': 'RAFT',
    'srv': 'SRV',
    'defender': 'Defender',
    'hoverquad': 'HoverQuad',
    'mustang': 'Mustang',
    'alpha': 'Alpha',
    'beta': 'Beta',
    'delta': 'Delta',
    'gamma': 'Gamma',
    'omega': 'Omega',
    'nomad': 'Nomad',
    'intrepid': 'Intrepid',
    'spirit': 'Spirit',
    'mercury': 'Mercury',
    'star': 'Star',
    'runner': 'Runner',
    'starfighter': 'Starfighter',
    'ares': 'Ares',
    'inferno': 'Inferno',
    'ion': 'Ion',
    'starlifter': 'Starlifter',
    'hercules': 'Hercules',
    'buccaneer': 'Buccaneer',
    'caterpillar': 'Caterpillar',
    'pirate': 'Pirate',
    'clipper': 'Clipper',
    'corsair': 'Corsair',
    'cutlass': 'Cutlass',
    'black': 'Black',
    'blue': 'Blue',
    'red': 'Red',
    'steel': 'Steel',
    'cutter': 'Cutter',
    'rambler': 'Rambler',
    'scout': 'Scout',
    'dragonfly': 'Dragonfly',
    'yellowjacket': 'Yellowjacket',
    'golem': 'Golem',
    'herald': 'Herald',
    'vulture': 'Vulture',
    'prowler': 'Prowler',
    'talon': 'Talon',
    'shrike': 'Shrike',
    'syulen': 'Syulen',
    'wolf': 'Wolf',
    'alphawolf': 'AlphaWolf',
    'merlin': 'Merlin',
    'archimedes': 'Archimedes',
    'emerald': 'Emerald',
    'fortune': 'Fortune',
    'freelancer': 'Freelancer',
    'dur': 'DUR',
    'max': 'MAX',
    'mis': 'MIS',
    'fury': 'Fury',
    'hull': 'Hull',
    'prospector': 'Prospector',
    'razor': 'Razor',
    'ex': 'EX',
    'lx': 'LX',
    'reliant': 'Reliant',
    'kore': 'Kore',
    'mako': 'Mako',
    'sen': 'Sen',
    'tana': 'Tana',
    'starfarer': 'Starfarer',
    'gemini': 'Gemini',
    'starlancer': 'Starlancer',
    'tac': 'TAC',
    'guardian': 'Guardian',
    'mx': 'MX',
    'qi': 'Qi',
    'pulse': 'Pulse',
    '100i': '100i',
    '125a': '125a',
    '135c': '135c',
    '300i': '300i',
    '315p': '315p',
    '325a': '325a',
    '350r': '350r',
    '400i': '400i',
    '600i': '600i',
    'touring': 'Touring',
    '85x': '85X',
    '890jump': '890 Jump',
    'm50': 'M50',
    'x1': 'X1',
    'force': 'Force',
    'velocity': 'Velocity',
    'apollo': 'Apollo',
    'medivac': 'Medivac',
    'triage': 'Triage',
    'aurora': 'Aurora',
    'cl': 'CL',
    'es': 'ES',
    'ln': 'LN',
    'mr': 'MR',
    'constellation': 'Constellation',
    'andromeda': 'Andromeda',
    'aquila': 'Aquila',
    'phoenix': 'Phoenix',
    'taurus': 'Taurus',
    'mantis': 'Mantis',
    'meteor': 'Meteor',
    'perseus': 'Perseus',
    'polaris': 'Polaris',
    'salvation': 'Salvation',
    'scorpius': 'Scorpius',
    'antares': 'Antares',
    'zeus': 'Zeus',
    'blade': 'Blade',
    'glaive': 'Glaive',
    'scythe': 'Scythe',
    'stinger': 'Stinger',
    'santokyai': "San'tok.yāi",
    'nox': 'Nox',
    'kue': 'Kue',
    'f7a': 'F7A',
    'f7c': 'F7C',
    'f7cm': 'F7C-M',
    'f7cr': 'F7C-R',
    'f7cs': 'F7C-S',
    'f8': 'F8',
    'f8c': 'F8C',
    'mk1': 'Mk I',
    'mk2': 'Mk II',
    'c8': 'C8',
    'c8r': 'C8R',
    'c8x': 'C8X',
    'a1': 'A1',
    'a2': 'A2',
    'c1': 'C1',
    'c2': 'C2',
    'm2': 'M2',
    'p52': 'P-52',
    'p72': 'P-72',
    'l21': 'L21',
    'l22': 'L22',
    'wildfire': 'Wildfire',
    'heartseeker': 'Heartseeker',
    'executive': 'Executive',
    'edition': 'Edition',
}

# Patterns to filter out (AI variants, special versions, etc.)
FILTER_PATTERNS = [
    r'_pu_ai_',        # PU AI variants
    r'_ai_',           # General AI variants
    r'_ea_ai_',        # EA AI variants
    r'ai_override',    # AI override
    r'_teach$',        # Teaching variants
    r'_tutorial$',     # Tutorial variants
    r'_hijacked',      # Hijacked versions
    r'_derelict',      # Derelict wrecks
    r'_wreck',         # Wrecks
    r'nodebris',       # No debris versions
    r'_fw_\d+',        # Fleetweek dated
    r'_fw\d+',         # Fleetweek
    r'fleetweek',      # Fleetweek
    r'indestructible', # Indestructible
    r'gamemaster',     # Gamemaster
    r'invictus',       # Invictus versions
    r'showdown',       # Showdown event
    r'citizencon',     # Citizencon
    r'_piano',         # Special versions
    r'_drug_',         # Drug running versions
    r'mission_pir',    # Mission variants
    r'crewless',       # Crewless
    r'_pir$',          # Pirate suffix (but not "pirate")
    r'_dunlevy$',      # Named variants
    r'_crocodile$',    # Named variants
    r'nointerior',     # No interior
    r'_pu$',           # PU suffix only
    r'bis\d{4}',       # BIS year versions
    r'_bis_',          # BIS versions
    r'collector_',     # Collector editions (handled separately)
    r'exec_',          # Executive editions (handled separately)
    r'_plat$',         # Platinum
    r'bombless',       # Bombless
    r's3bombs',        # Special bomb versions
    r'_civilian$',     # Civilian suffix
    r'tier_\d',        # Tier versions
    r'_pink$',         # Color variants
    r'_yellow$',       # Color variants
]

# Styles
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CALC_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
RESULT_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
ATTACKER_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
TARGET_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SKILL_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
TTK_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
ARMOR_FILL = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
THRUSTER_FILL = PatternFill(start_color="AED6F1", end_color="AED6F1", fill_type="solid")
COMPONENT_FILL = PatternFill(start_color="E8F6F3", end_color="E8F6F3", fill_type="solid")
ZONE_FILL = PatternFill(start_color="F5EEF8", end_color="F5EEF8", fill_type="solid")
POWER_FILL = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SECTION_FONT = Font(bold=True, size=12, color="2E75B6")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

MOUNT_TYPES = {'Fixed': 0.70, 'Gimballed': 0.85, 'Turret': 0.90, 'Auto-Gimbal': 0.95}
FIRE_MODES = {'Sustained': 1.00, 'Burst': 0.85, 'Staggered': 0.75}
POWER_TRIANGLE = {0.00: 1.00, 0.33: 1.00, 0.50: 1.07, 0.66: 1.13, 1.00: 1.20}
COMBAT_SCENARIOS = {
    'Dogfight': (0.75, 0.65),
    'Synthetic Test': (0.95, 0.95),
    'Jousting': (0.85, 0.35),
    'Custom': (0.80, 0.80),
}
TARGET_ZONES = {
    'Center Mass': {'hull': 0.6, 'armor': 0.3, 'thrusters': 0.05, 'components': 0.05},
    'Engines': {'hull': 0.2, 'armor': 0.1, 'thrusters': 0.6, 'components': 0.1},
    'Cockpit': {'hull': 0.5, 'armor': 0.2, 'thrusters': 0.0, 'components': 0.3},
    'Wings/Extremities': {'hull': 0.3, 'armor': 0.4, 'thrusters': 0.2, 'components': 0.1},
    'Turrets': {'hull': 0.1, 'armor': 0.1, 'thrusters': 0.0, 'components': 0.0},
}


def load_csv(filepath):
    if not filepath.exists():
        print(f"  Warning: File not found: {filepath}")
        return []
    with open(filepath, 'r', encoding='utf-8') as f:
        return list(csv.DictReader(f))


def should_filter_ship(filename):
    """Check if this ship should be filtered out."""
    filename_lower = filename.lower()
    for pattern in FILTER_PATTERNS:
        if re.search(pattern, filename_lower):
            return True
    return False


def format_ship_name(filename):
    """Create a properly formatted display name from filename."""
    # Remove manufacturer prefix and split into parts
    parts = filename.lower().split('_')
    if len(parts) < 2:
        return filename

    mfr_code = parts[0].upper()
    mfr_name = MANUFACTURERS.get(mfr_code, mfr_code)

    # Build the model name from remaining parts
    model_parts = []
    for part in parts[1:]:
        # Check for known name fixes
        if part in SHIP_NAME_FIXES:
            model_parts.append(SHIP_NAME_FIXES[part])
        else:
            # Title case unknown parts
            model_parts.append(part.title())

    model_name = ' '.join(model_parts)

    # Clean up common issues
    model_name = model_name.replace('  ', ' ')

    return f"{mfr_name} {model_name}"


def set_column_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def apply_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER


def apply_attacker(cell):
    cell.fill = ATTACKER_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='left')
    cell.font = Font(bold=True)


def apply_target(cell):
    cell.fill = TARGET_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='left')
    cell.font = Font(bold=True)


def apply_skill(cell):
    cell.fill = SKILL_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')


def apply_calc(cell):
    cell.fill = CALC_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')


def apply_result(cell):
    cell.fill = RESULT_FILL
    cell.border = THIN_BORDER
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')


def apply_armor(cell):
    cell.fill = ARMOR_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')


def apply_thruster(cell):
    cell.fill = THRUSTER_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')


def apply_component(cell):
    cell.fill = COMPONENT_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')


def apply_zone(cell):
    cell.fill = ZONE_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')


def apply_power(cell):
    cell.fill = POWER_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')


def create_weapons_data_sheet(wb):
    """Create weapons data sheet with proper naming."""
    ws = wb.create_sheet("_WeaponsData")
    weapons_data = load_csv(DATA_ROOT / "weapons" / "ship_weapons.csv")
    power_data = load_csv(DATA_ROOT / "weapons" / "weapon_power_data.csv")

    power_lookup = {}
    for p in power_data:
        fname = p.get('filename', '').lower()
        power_lookup[fname] = float(p.get('power_consumption', 0) or 0)

    if not weapons_data:
        return {}, 0

    weapons_by_size = {i: [] for i in range(1, 10)}

    for w in weapons_data:
        size = int(w.get('size', 0))
        if 1 <= size <= 9:
            display = w.get('display_name', w.get('filename', 'Unknown'))
            filename = w.get('filename', '')
            damage_type = w.get('damage_type', 'Unknown')
            sustained_dps = float(w.get('sustained_dps', 0) or w.get('dps', 0) or 0)
            power_consumption = power_lookup.get(filename.lower(), 0)

            weapons_by_size[size].append({
                'display': display,
                'filename': filename,
                'damage_type': damage_type,
                'sustained_dps': sustained_dps,
                'size': size,
                'power_consumption': power_consumption,
            })

    for size in weapons_by_size:
        weapons_by_size[size].sort(key=lambda x: (-x['sustained_dps'], x['display']))

    for size in range(1, 10):
        col = size
        ws.cell(row=1, column=col, value=f"S{size} Weapons")
        apply_header(ws.cell(row=1, column=col))
        for row, w in enumerate(weapons_by_size[size], 2):
            ws.cell(row=row, column=col, value=w['display'])

    ws.cell(row=1, column=11, value="Weapon Name")
    ws.cell(row=1, column=12, value="Damage Type")
    ws.cell(row=1, column=13, value="Size")
    ws.cell(row=1, column=14, value="Sustained DPS")
    ws.cell(row=1, column=15, value="Power Consumption")

    all_weapons = []
    for size, weapons in weapons_by_size.items():
        for w in weapons:
            all_weapons.append(w)
    all_weapons.sort(key=lambda x: (x['size'], -x['sustained_dps'], x['display']))

    for row, w in enumerate(all_weapons, 2):
        ws.cell(row=row, column=11, value=w['display'])
        ws.cell(row=row, column=12, value=w['damage_type'])
        ws.cell(row=row, column=13, value=w['size'])
        ws.cell(row=row, column=14, value=w['sustained_dps'])
        ws.cell(row=row, column=15, value=w['power_consumption'])

    ws.sheet_state = 'hidden'
    size_counts = {size: len(weapons_by_size[size]) for size in range(1, 10)}
    return size_counts, len(all_weapons)


def create_ships_data_sheet(wb):
    """Create ships data sheet with FILTERED player ships and proper names."""
    ws = wb.create_sheet("_ShipsData")

    parts_data = load_csv(DATA_ROOT / "ships" / "ship_parts_comprehensive.csv")
    hardpoint_data = load_csv(DATA_ROOT / "ships" / "ship_hardpoint_summary.csv")
    hull_data = load_csv(DATA_ROOT / "ships" / "ship_hull_hp.csv")

    hardpoint_lookup = {}
    for hp in hardpoint_data:
        ship_name = hp.get('ship_name', '').upper()
        hardpoint_lookup[ship_name] = hp

    hull_lookup = {}
    for h in hull_data:
        display = h.get('display_name', '')
        internal = h.get('internal_name', '').upper()
        filename = h.get('filename', '').upper()
        hull_info = {
            'hull_hp': int(h.get('hull_hp', 0) or 0),
            'total_shield_hp': float(h.get('total_shield_hp', 0) or 0),
            'total_shield_regen': float(h.get('total_shield_regen', 0) or 0),
            'shield_count': int(h.get('shield_count', 0) or 0),
            'shield_size': int(h.get('shield_size', 1) or 1),
            'shield_ref': h.get('shield_ref', ''),
        }
        if display:
            hull_lookup[display.upper()] = hull_info
        if internal:
            hull_lookup[internal] = hull_info
        if filename:
            hull_lookup[filename] = hull_info

    ships = []
    filtered_count = 0

    for part in parts_data:
        filename = part.get('filename', '')

        # Filter out AI variants and special versions
        if should_filter_ship(filename):
            filtered_count += 1
            continue

        ship_upper = filename.upper()
        hp_info = hardpoint_lookup.get(ship_upper, {})

        pilot_count = int(hp_info.get('pilot_weapon_count', 0) or 0)
        if pilot_count == 0:
            continue

        hull_info = hull_lookup.get(ship_upper) or {}

        # Create proper display name
        display_name = format_ship_name(filename)

        ships.append({
            'display_name': display_name,
            'filename': filename,
            'hull_hp_normalized': int(part.get('hull_hp_normalized', 0) or 0),
            'fuse_penetration_mult': float(part.get('fuse_penetration_mult', 1) or 1),
            'component_penetration_mult': float(part.get('component_penetration_mult', 1) or 1),
            'critical_explosion_chance': float(part.get('critical_explosion_chance', 0.2) or 0.2),
            'armor_hp': int(part.get('armor_hp', 0) or 0),
            'armor_resist_physical': float(part.get('armor_resist_physical', 1) or 1),
            'armor_resist_energy': float(part.get('armor_resist_energy', 1) or 1),
            'armor_resist_distortion': float(part.get('armor_resist_distortion', 1) or 1),
            'thruster_count': int(part.get('thruster_count', 0) or 0),
            'thruster_main_hp': int(part.get('thruster_main_hp', 0) or 0),
            'thruster_retro_hp': int(part.get('thruster_retro_hp', 0) or 0),
            'thruster_mav_hp': int(part.get('thruster_mav_hp', 0) or 0),
            'thruster_vtol_hp': int(part.get('thruster_vtol_hp', 0) or 0),
            'thruster_total_hp': int(part.get('thruster_total_hp', 0) or 0),
            'turret_count': int(part.get('turret_count', 0) or 0),
            'turret_total_hp': int(part.get('turret_total_hp', 0) or 0),
            'powerplant_count': int(part.get('powerplant_count', 0) or 0),
            'powerplant_total_hp': int(part.get('powerplant_total_hp', 0) or 0),
            'cooler_count': int(part.get('cooler_count', 0) or 0),
            'cooler_total_hp': int(part.get('cooler_total_hp', 0) or 0),
            'shield_gen_count': int(part.get('shield_gen_count', 0) or 0),
            'shield_gen_total_hp': int(part.get('shield_gen_total_hp', 0) or 0),
            'qd_count': int(part.get('qd_count', 0) or 0),
            'qd_total_hp': int(part.get('qd_total_hp', 0) or 0),
            'total_shield_hp': hull_info.get('total_shield_hp', 0),
            'total_shield_regen': hull_info.get('total_shield_regen', 0),
            'shield_size': hull_info.get('shield_size', 1),
            'shield_count': hull_info.get('shield_count', 0),
            'shield_ref': hull_info.get('shield_ref', ''),
            'pilot_weapon_count': pilot_count,
            'pilot_weapon_sizes': hp_info.get('pilot_weapon_sizes', ''),
        })

    ships.sort(key=lambda x: x['display_name'])
    print(f"  Filtered out {filtered_count} AI/special variants")

    headers = [
        "Name", "Hull HP Norm", "Fuse Pen Mult", "Comp Pen Mult", "Crit Chance",
        "Armor HP", "Armor Phys", "Armor Energy", "Armor Dist",
        "Thruster Count", "Thruster Main", "Thruster Retro", "Thruster Mav", "Thruster VTOL", "Thruster Total",
        "Turret Count", "Turret HP",
        "PP Count", "PP HP", "Cooler Count", "Cooler HP", "Shield Gen Count", "Shield Gen HP", "QD Count", "QD HP",
        "Shield HP", "Shield Regen", "Shield Size", "Shield Count", "Shield Ref",
        "Pilot Weapons", "Pilot Sizes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for row, s in enumerate(ships, 2):
        ws.cell(row=row, column=1, value=s['display_name'])
        ws.cell(row=row, column=2, value=s['hull_hp_normalized'])
        ws.cell(row=row, column=3, value=s['fuse_penetration_mult'])
        ws.cell(row=row, column=4, value=s['component_penetration_mult'])
        ws.cell(row=row, column=5, value=s['critical_explosion_chance'])
        ws.cell(row=row, column=6, value=s['armor_hp'])
        ws.cell(row=row, column=7, value=s['armor_resist_physical'])
        ws.cell(row=row, column=8, value=s['armor_resist_energy'])
        ws.cell(row=row, column=9, value=s['armor_resist_distortion'])
        ws.cell(row=row, column=10, value=s['thruster_count'])
        ws.cell(row=row, column=11, value=s['thruster_main_hp'])
        ws.cell(row=row, column=12, value=s['thruster_retro_hp'])
        ws.cell(row=row, column=13, value=s['thruster_mav_hp'])
        ws.cell(row=row, column=14, value=s['thruster_vtol_hp'])
        ws.cell(row=row, column=15, value=s['thruster_total_hp'])
        ws.cell(row=row, column=16, value=s['turret_count'])
        ws.cell(row=row, column=17, value=s['turret_total_hp'])
        ws.cell(row=row, column=18, value=s['powerplant_count'])
        ws.cell(row=row, column=19, value=s['powerplant_total_hp'])
        ws.cell(row=row, column=20, value=s['cooler_count'])
        ws.cell(row=row, column=21, value=s['cooler_total_hp'])
        ws.cell(row=row, column=22, value=s['shield_gen_count'])
        ws.cell(row=row, column=23, value=s['shield_gen_total_hp'])
        ws.cell(row=row, column=24, value=s['qd_count'])
        ws.cell(row=row, column=25, value=s['qd_total_hp'])
        ws.cell(row=row, column=26, value=s['total_shield_hp'])
        ws.cell(row=row, column=27, value=s['total_shield_regen'])
        ws.cell(row=row, column=28, value=s['shield_size'])
        ws.cell(row=row, column=29, value=s['shield_count'])
        ws.cell(row=row, column=30, value=s['shield_ref'])
        ws.cell(row=row, column=31, value=s['pilot_weapon_count'])
        ws.cell(row=row, column=32, value=s['pilot_weapon_sizes'])

    ws.sheet_state = 'hidden'
    return [s['display_name'] for s in ships], len(ships)


def create_shields_data_sheet(wb):
    """Create shields data sheet with proper naming."""
    ws = wb.create_sheet("_ShieldsData")
    shields_data = load_csv(DATA_ROOT / "components" / "shields" / "shields.csv")

    # Manufacturer codes -> full names
    SHIELD_MFR = {
        'GODI': 'Gorgon Defender', 'ASAS': 'ASAS', 'BASL': 'Basilisk',
        'SECO': 'Seal Corp', 'BANU': 'Banu', 'BEHR': 'Behring', 'RSI': 'RSI',
        'SUKS': 'Sukoran', 'YORM': 'Yorm',
    }

    # Model name formatting
    SHIELD_MODELS = {
        'allstop': 'AllStop', 'forcewall': 'ForceWall', 'fr66': 'FR-66',
        'securehyde': 'SecureHyde', 'bulwark': 'Bulwark', 'guardian': 'Guardian',
        'palisade': 'Palisade', 'steward': 'Steward', '5sa': '5SA', '6sa': '6SA',
        '7sa': '7SA', 'castra': 'Castra', 'pin': 'Pin', 'hex': 'Hex', 'ink': 'Ink',
        'web': 'Web', 'cloak': 'Cloak', 'mirage': 'Mirage', 'shimmer': 'Shimmer',
        'veil': 'Veil', 'placeholder': 'Standard', 'falco': 'Falco',
        'centurion': 'Centurion', 'umbra': 'Umbra', 'siren': 'Siren',
        'resistgasclouds': '(Gas Resist)',
    }

    shields = []
    shields_by_size = {i: [] for i in range(0, 5)}

    for shield in shields_data:
        name = shield.get('name', '')
        size = int(shield.get('size', 0) or 0)
        max_hp = float(shield.get('shield_max_hp', 0) or 0)

        if 'Template' in name or max_hp <= 0 or size < 0:
            continue

        # Parse shield name: shld_MFGR_sXX_MODEL_scitem
        parts = name.lower().replace('_scitem', '').split('_')

        # Standard format: shld_mfgr_sXX_model[_extra]
        if len(parts) >= 4 and parts[0] == 'shld':
            mfr_code = parts[1].upper()
            mfr_name = SHIELD_MFR.get(mfr_code, mfr_code.title())

            # Model is everything after size code
            model_parts = []
            for part in parts[3:]:
                if part in SHIELD_MODELS:
                    model_parts.append(SHIELD_MODELS[part])
                else:
                    model_parts.append(part.upper() if len(part) <= 3 else part.title())

            model = ' '.join(model_parts)
        else:
            # Fallback
            mfr_name = parts[1].title() if len(parts) > 1 else 'Unknown'
            model = parts[-1].title() if len(parts) > 2 else name

        display_name = f"{mfr_name} {model}".strip()

        shield_info = {
            'display_name': display_name,
            'internal_name': name,
            'size': size,
            'max_hp': max_hp,
            'regen': float(shield.get('shield_regen', 0) or 0),
            'resist_physical_avg': float(shield.get('resist_physical_avg', 0) or 0),
            'resist_energy_avg': float(shield.get('resist_energy_avg', 0) or 0),
            'resist_distortion_avg': float(shield.get('resist_distortion_avg', 0) or 0),
            'absorb_physical_avg': float(shield.get('absorb_physical_avg', 0) or 0),
            'absorb_energy_avg': float(shield.get('absorb_energy_avg', 0) or 0),
            'absorb_distortion_avg': float(shield.get('absorb_distortion_avg', 0) or 0),
        }

        shields.append(shield_info)
        if 0 <= size <= 4:
            shields_by_size[size].append(shield_info)

    shields.sort(key=lambda x: (x['size'], -x['max_hp']))
    for size in shields_by_size:
        shields_by_size[size].sort(key=lambda x: -x['max_hp'])

    for size in range(0, 5):
        col = size + 1
        ws.cell(row=1, column=col, value=f"S{size} Shields")
        apply_header(ws.cell(row=1, column=col))
        for row, s in enumerate(shields_by_size[size], 2):
            ws.cell(row=row, column=col, value=s['display_name'])

    headers = ["Display Name", "Internal Name", "Size", "Max HP", "Regen/sec",
               "Resist Phys", "Resist Energy", "Resist Dist",
               "Absorb Phys", "Absorb Energy", "Absorb Dist"]
    for col, h in enumerate(headers, 7):
        ws.cell(row=1, column=col, value=h)
        apply_header(ws.cell(row=1, column=col))

    for row, s in enumerate(shields, 2):
        ws.cell(row=row, column=7, value=s['display_name'])
        ws.cell(row=row, column=8, value=s['internal_name'])
        ws.cell(row=row, column=9, value=s['size'])
        ws.cell(row=row, column=10, value=s['max_hp'])
        ws.cell(row=row, column=11, value=s['regen'])
        ws.cell(row=row, column=12, value=s['resist_physical_avg'])
        ws.cell(row=row, column=13, value=s['resist_energy_avg'])
        ws.cell(row=row, column=14, value=s['resist_distortion_avg'])
        ws.cell(row=row, column=15, value=s['absorb_physical_avg'])
        ws.cell(row=row, column=16, value=s['absorb_energy_avg'])
        ws.cell(row=row, column=17, value=s['absorb_distortion_avg'])

    ws.sheet_state = 'hidden'
    shield_size_counts = {size: len(shields_by_size[size]) for size in range(0, 5)}
    return [s['display_name'] for s in shields], [s['internal_name'] for s in shields], len(shields), shield_size_counts


def create_modifiers_sheet(wb):
    ws = wb.create_sheet("_Modifiers")

    ws['A1'] = "Mount Type"
    ws['B1'] = "Accuracy Mod"
    for row, (mt, mod) in enumerate(MOUNT_TYPES.items(), 2):
        ws.cell(row=row, column=1, value=mt)
        ws.cell(row=row, column=2, value=mod)

    ws['D1'] = "Fire Mode"
    ws['E1'] = "DPS Mod"
    for row, (fm, mod) in enumerate(FIRE_MODES.items(), 2):
        ws.cell(row=row, column=4, value=fm)
        ws.cell(row=row, column=5, value=mod)

    ws['G1'] = "Power Level"
    ws['H1'] = "Power Mult"
    for row, (pwr, mod) in enumerate(POWER_TRIANGLE.items(), 2):
        ws.cell(row=row, column=7, value=pwr)
        ws.cell(row=row, column=8, value=mod)

    ws['J1'] = "Scenario"
    ws['K1'] = "Accuracy"
    ws['L1'] = "ToT"
    for row, (scenario, (acc, tot)) in enumerate(COMBAT_SCENARIOS.items(), 2):
        ws.cell(row=row, column=10, value=scenario)
        ws.cell(row=row, column=11, value=acc)
        ws.cell(row=row, column=12, value=tot)

    ws['N1'] = "Zone"
    ws['O1'] = "Hull %"
    ws['P1'] = "Armor %"
    ws['Q1'] = "Thruster %"
    ws['R1'] = "Component %"
    for row, (zone, dist) in enumerate(TARGET_ZONES.items(), 2):
        ws.cell(row=row, column=14, value=zone)
        ws.cell(row=row, column=15, value=dist['hull'])
        ws.cell(row=row, column=16, value=dist['armor'])
        ws.cell(row=row, column=17, value=dist['thrusters'])
        ws.cell(row=row, column=18, value=dist['components'])

    ws.sheet_state = 'hidden'


def create_named_ranges(wb, size_counts, total_weapons, shield_size_counts, total_shields):
    for size in range(1, 10):
        count = size_counts.get(size, 0)
        if count > 0:
            range_name = f"Weapons_S{size}"
            range_ref = f"_WeaponsData!${get_column_letter(size)}$2:${get_column_letter(size)}${count + 1}"
            wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(range_name, attr_text=range_ref))

    all_range_ref = f"_WeaponsData!$K$2:$K${total_weapons + 1}"
    wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName("Weapons_All", attr_text=all_range_ref))

    for size in range(0, 5):
        count = shield_size_counts.get(size, 0)
        if count > 0:
            range_name = f"Shields_S{size}"
            col = size + 1
            range_ref = f"_ShieldsData!${get_column_letter(col)}$2:${get_column_letter(col)}${count + 1}"
            wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(range_name, attr_text=range_ref))

    all_shields_ref = f"_ShieldsData!$G$2:$G${total_shields + 1}"
    wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName("Shields_All", attr_text=all_shields_ref))


def create_calculator_sheet(wb, ship_names, shield_names, shield_internal_names, ship_count, shield_count, size_counts, shield_size_counts):
    """Create the main calculator sheet."""
    ws = wb.create_sheet("Calculator")

    ws['A1'] = "STAR CITIZEN DAMAGE CALCULATOR v12"
    ws['A1'].font = Font(bold=True, size=18, color="1F4E79")
    ws.merge_cells('A1:G1')

    ws['A2'] = "Full Ship Parts System - Armor, Thrusters, Components"
    ws['A2'].font = Font(italic=True, color="666666")

    ws['I1'] = "TARGET PARTS BREAKDOWN"
    ws['I1'].font = Font(bold=True, size=14, color="C0392B")
    ws.merge_cells('I1:L1')

    row = 4

    # SECTION 1: YOUR SHIP
    ws.cell(row=row, column=1, value="STEP 1: SELECT YOUR SHIP")
    ws.cell(row=row, column=1).font = TITLE_FONT
    row += 2

    ws.cell(row=row, column=2, value="Your Ship:")
    ws.cell(row=row, column=2).font = Font(bold=True)
    ship_cell = ws.cell(row=row, column=3, value=ship_names[0] if ship_names else "")
    apply_attacker(ship_cell)
    your_ship_row = row

    ship_dv = DataValidation(type="list", formula1=f"_ShipsData!$A$2:$A${ship_count+1}", allow_blank=True)
    ws.add_data_validation(ship_dv)
    ship_dv.add(ship_cell)
    row += 2

    ws.cell(row=row, column=2, value="Weapon Sizes:")
    ws.cell(row=row, column=3, value=f'=IFERROR(VLOOKUP(C{your_ship_row},_ShipsData!A:AF,32,FALSE),"")')
    apply_calc(ws.cell(row=row, column=3))
    weapon_sizes_row = row
    row += 2

    # SECTION 2: COMBAT SCENARIO
    ws.cell(row=row, column=1, value="STEP 2: COMBAT SCENARIO")
    ws.cell(row=row, column=1).font = TITLE_FONT
    row += 2

    ws.cell(row=row, column=2, value="Scenario:")
    scenario_cell = ws.cell(row=row, column=3, value="Dogfight")
    apply_skill(scenario_cell)
    scenario_cell.font = Font(bold=True)
    scenario_row = row
    scenario_dv = DataValidation(type="list", formula1='"Dogfight,Synthetic Test,Jousting,Custom"', allow_blank=False)
    ws.add_data_validation(scenario_dv)
    scenario_dv.add(scenario_cell)
    row += 2

    ws.cell(row=row, column=2, value="Pilot Accuracy:")
    acc_cell = ws.cell(row=row, column=3, value=f'=VLOOKUP(C{scenario_row},_Modifiers!J:L,2,FALSE)')
    acc_cell.number_format = '0%'
    apply_skill(acc_cell)
    accuracy_row = row
    row += 1

    ws.cell(row=row, column=2, value="Mount Type:")
    mount_cell = ws.cell(row=row, column=3, value="Gimballed")
    apply_skill(mount_cell)
    mount_row = row
    mount_dv = DataValidation(type="list", formula1='"Fixed,Gimballed,Turret,Auto-Gimbal"', allow_blank=False)
    ws.add_data_validation(mount_dv)
    mount_dv.add(mount_cell)
    row += 1

    ws.cell(row=row, column=2, value="Fire Mode:")
    fire_cell = ws.cell(row=row, column=3, value="Sustained")
    apply_skill(fire_cell)
    fire_row = row
    fire_dv = DataValidation(type="list", formula1='"Sustained,Burst,Staggered"', allow_blank=False)
    ws.add_data_validation(fire_dv)
    fire_dv.add(fire_cell)
    row += 1

    ws.cell(row=row, column=2, value="Time-on-Target:")
    tot_cell = ws.cell(row=row, column=3, value=f'=VLOOKUP(C{scenario_row},_Modifiers!J:L,3,FALSE)')
    tot_cell.number_format = '0%'
    apply_skill(tot_cell)
    tot_row = row
    row += 1

    ws.cell(row=row, column=2, value="Weapon Power:")
    power_cell = ws.cell(row=row, column=3, value=0.33)
    apply_power(power_cell)
    power_cell.number_format = '0%'
    power_row = row
    power_dv = DataValidation(type="list", formula1='"0,0.33,0.5,0.66,1"', allow_blank=False)
    ws.add_data_validation(power_dv)
    power_dv.add(power_cell)
    row += 2

    ws.cell(row=row, column=2, value="Mount Bonus:")
    ws.cell(row=row, column=3, value=f'=VLOOKUP(C{mount_row},_Modifiers!A:B,2,FALSE)')
    apply_calc(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).number_format = '0%'
    mount_bonus_row = row
    row += 1

    ws.cell(row=row, column=2, value="Fire Mode Factor:")
    ws.cell(row=row, column=3, value=f'=VLOOKUP(C{fire_row},_Modifiers!D:E,2,FALSE)')
    apply_calc(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).number_format = '0%'
    fire_factor_row = row
    row += 1

    ws.cell(row=row, column=2, value="Power Multiplier:")
    ws.cell(row=row, column=3, value=f'=VLOOKUP(C{power_row},_Modifiers!G:H,2,FALSE)')
    apply_power(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).number_format = '0%'
    power_mult_row = row
    row += 1

    ws.cell(row=row, column=2, value="Hit Rate:")
    ws.cell(row=row, column=3, value=f'=MIN(C{accuracy_row}*C{mount_bonus_row},1)')
    apply_result(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).number_format = '0%'
    hit_rate_row = row
    row += 2

    # SECTION 3: WEAPONS
    ws.cell(row=row, column=1, value="STEP 3: WEAPONS")
    ws.cell(row=row, column=1).font = TITLE_FONT
    row += 2

    slot_headers = ['Slot', 'Size', 'Weapon', 'Type', 'Sust DPS', 'Power', 'Eff DPS']
    for col, h in enumerate(slot_headers, 2):
        cell = ws.cell(row=row, column=col, value=h)
        apply_header(cell)
    row += 1

    slot_start_row = row

    for slot in range(1, MAX_WEAPON_SLOTS + 1):
        ws.cell(row=row, column=2, value=slot)
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')

        size_formula = f'=IFERROR(VALUE(TRIM(MID(SUBSTITUTE($C${weapon_sizes_row},",",REPT(" ",100)),({slot}-1)*100+1,100))),"")'
        size_cell = ws.cell(row=row, column=3, value=size_formula)
        apply_calc(size_cell)

        weapon_cell = ws.cell(row=row, column=4, value="")
        apply_attacker(weapon_cell)

        indirect_formula = f'INDIRECT(IF(C{row}="","Weapons_All","Weapons_S"&C{row}))'
        weapon_dv = DataValidation(type="list", formula1=indirect_formula, allow_blank=True)
        ws.add_data_validation(weapon_dv)
        weapon_dv.add(weapon_cell)

        ws.cell(row=row, column=5, value=f'=IFERROR(VLOOKUP(D{row},_WeaponsData!$K$2:$O$200,2,FALSE),"")')
        apply_calc(ws.cell(row=row, column=5))

        ws.cell(row=row, column=6, value=f'=IFERROR(VLOOKUP(D{row},_WeaponsData!$K$2:$O$200,4,FALSE),0)')
        apply_calc(ws.cell(row=row, column=6))
        ws.cell(row=row, column=6).number_format = '#,##0'

        ws.cell(row=row, column=7, value=f'=IFERROR(VLOOKUP(D{row},_WeaponsData!$K$2:$O$200,5,FALSE),0)')
        apply_power(ws.cell(row=row, column=7))
        ws.cell(row=row, column=7).number_format = '0.00'

        dps_cell = ws.cell(row=row, column=8, value=f'=F{row}*$C${hit_rate_row}*$C${fire_factor_row}*$C${tot_row}*$C${power_mult_row}')
        apply_calc(dps_cell)
        dps_cell.number_format = '#,##0'

        row += 1

    slot_end_row = row - 1

    row += 1
    ws.cell(row=row, column=2, value="TOTALS:")
    ws.cell(row=row, column=2).font = Font(bold=True)

    total_dps_cell = ws.cell(row=row, column=8, value=f'=SUM(H{slot_start_row}:H{slot_end_row})')
    apply_result(total_dps_cell)
    total_dps_cell.number_format = '#,##0'
    total_dps_cell.font = Font(bold=True, size=14)
    total_dps_row = row
    row += 2

    # SECTION 4: TARGET
    ws.cell(row=row, column=1, value="STEP 4: TARGET")
    ws.cell(row=row, column=1).font = TITLE_FONT
    row += 2

    ws.cell(row=row, column=2, value="Target Ship:")
    target_cell = ws.cell(row=row, column=3, value=ship_names[0] if ship_names else "")
    apply_target(target_cell)
    target_ship_row = row
    target_dv = DataValidation(type="list", formula1=f"_ShipsData!$A$2:$A${ship_count+1}", allow_blank=True)
    ws.add_data_validation(target_dv)
    target_dv.add(target_cell)
    row += 1

    ws.cell(row=row, column=2, value="Target Zone:")
    ws.cell(row=row, column=2).font = Font(bold=True, color="C0392B")
    zone_cell = ws.cell(row=row, column=3, value="Center Mass")
    apply_zone(zone_cell)
    zone_cell.font = Font(bold=True)
    target_zone_row = row
    zone_dv = DataValidation(type="list", formula1='"Center Mass,Engines,Cockpit,Wings/Extremities,Turrets"', allow_blank=False)
    ws.add_data_validation(zone_dv)
    zone_dv.add(zone_cell)
    row += 1

    ws.cell(row=row, column=2, value="Shield Size:")
    ws.cell(row=row, column=3, value=f'=IFERROR(VLOOKUP(C{target_ship_row},_ShipsData!A:AF,28,FALSE),1)')
    apply_calc(ws.cell(row=row, column=3))
    target_shield_size_row = row
    row += 1

    ws.cell(row=row, column=2, value="Target Shield:")
    shield_cell = ws.cell(row=row, column=3, value=shield_names[0] if shield_names else "")
    apply_target(shield_cell)
    target_shield_row = row
    shield_dv = DataValidation(type="list", formula1=f'INDIRECT("Shields_S"&C{target_shield_size_row})', allow_blank=True)
    ws.add_data_validation(shield_dv)
    shield_dv.add(shield_cell)
    row += 2

    # SIDE PANEL
    side_row = 4

    ws.cell(row=side_row, column=9, value="HULL & ARMOR")
    ws.cell(row=side_row, column=9).font = Font(bold=True, size=11, color="C0392B")
    side_row += 1

    ws.cell(row=side_row, column=9, value="Hull HP:")
    ws.cell(row=side_row, column=10, value=f'=IFERROR(VLOOKUP($C${target_ship_row},_ShipsData!A:AF,2,FALSE),0)')
    apply_calc(ws.cell(row=side_row, column=10))
    ws.cell(row=side_row, column=10).number_format = '#,##0'
    hull_hp_row = side_row
    side_row += 1

    ws.cell(row=side_row, column=9, value="Armor HP:")
    ws.cell(row=side_row, column=10, value=f'=IFERROR(VLOOKUP($C${target_ship_row},_ShipsData!A:AF,6,FALSE),0)')
    apply_armor(ws.cell(row=side_row, column=10))
    ws.cell(row=side_row, column=10).number_format = '#,##0'
    armor_hp_row = side_row
    side_row += 2

    ws.cell(row=side_row, column=9, value="THRUSTERS")
    ws.cell(row=side_row, column=9).font = Font(bold=True, size=11, color="2980B9")
    side_row += 1

    ws.cell(row=side_row, column=9, value="Total HP:")
    ws.cell(row=side_row, column=10, value=f'=IFERROR(VLOOKUP($C${target_ship_row},_ShipsData!A:AF,15,FALSE),0)')
    apply_thruster(ws.cell(row=side_row, column=10))
    ws.cell(row=side_row, column=10).number_format = '#,##0'
    thruster_total_row = side_row
    side_row += 2

    ws.cell(row=side_row, column=9, value="COMPONENTS")
    ws.cell(row=side_row, column=9).font = Font(bold=True, size=11, color="27AE60")
    side_row += 1

    ws.cell(row=side_row, column=9, value="Power Plants:")
    ws.cell(row=side_row, column=10, value=f'=IFERROR(VLOOKUP($C${target_ship_row},_ShipsData!A:AF,19,FALSE),0)')
    apply_component(ws.cell(row=side_row, column=10))
    ws.cell(row=side_row, column=10).number_format = '#,##0'
    pp_hp_row = side_row
    side_row += 1

    ws.cell(row=side_row, column=9, value="Coolers:")
    ws.cell(row=side_row, column=10, value=f'=IFERROR(VLOOKUP($C${target_ship_row},_ShipsData!A:AF,21,FALSE),0)')
    apply_component(ws.cell(row=side_row, column=10))
    ws.cell(row=side_row, column=10).number_format = '#,##0'
    cooler_hp_row = side_row
    side_row += 1

    ws.cell(row=side_row, column=9, value="Shield Gens:")
    ws.cell(row=side_row, column=10, value=f'=IFERROR(VLOOKUP($C${target_ship_row},_ShipsData!A:AF,23,FALSE),0)')
    apply_component(ws.cell(row=side_row, column=10))
    ws.cell(row=side_row, column=10).number_format = '#,##0'
    shld_hp_row = side_row
    side_row += 2

    ws.cell(row=side_row, column=9, value="SHIELDS")
    ws.cell(row=side_row, column=9).font = Font(bold=True, size=11, color="3498DB")
    side_row += 1

    ws.cell(row=side_row, column=9, value="Shield HP:")
    ws.cell(row=side_row, column=10, value=f'=IFERROR(VLOOKUP(C{target_shield_row},_ShieldsData!$G:$K,4,FALSE)*VLOOKUP($C${target_ship_row},_ShipsData!A:AF,29,FALSE),0)')
    apply_calc(ws.cell(row=side_row, column=10))
    ws.cell(row=side_row, column=10).number_format = '#,##0'
    total_shield_hp_row = side_row
    side_row += 1

    ws.cell(row=side_row, column=9, value="Shield Regen:")
    ws.cell(row=side_row, column=10, value=f'=IFERROR(VLOOKUP(C{target_shield_row},_ShieldsData!$G:$K,5,FALSE)*VLOOKUP($C${target_ship_row},_ShipsData!A:AF,29,FALSE),0)')
    apply_calc(ws.cell(row=side_row, column=10))
    ws.cell(row=side_row, column=10).number_format = '#,##0'
    shield_regen_row = side_row

    # DAMAGE CALCULATIONS
    ws.cell(row=row, column=1, value="DAMAGE CALCULATIONS")
    ws.cell(row=row, column=1).font = TITLE_FONT
    row += 2

    ws.cell(row=row, column=2, value="SHIELD PHASE")
    ws.cell(row=row, column=2).font = SECTION_FONT
    row += 1

    ws.cell(row=row, column=2, value="DPS vs Shield:")
    ws.cell(row=row, column=3, value=f'=H{total_dps_row}')
    apply_calc(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).number_format = '#,##0'
    dps_vs_shield_row = row
    row += 1

    ws.cell(row=row, column=2, value="Net DPS (- regen):")
    ws.cell(row=row, column=3, value=f'=MAX(0,C{dps_vs_shield_row}-J{shield_regen_row})')
    apply_calc(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).number_format = '#,##0'
    net_shield_dps_row = row
    row += 1

    ws.cell(row=row, column=2, value="Time to Break Shield:")
    ws.cell(row=row, column=3, value=f'=IF(C{net_shield_dps_row}>0,J{total_shield_hp_row}/C{net_shield_dps_row},9999)')
    apply_result(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).number_format = '0.0'
    ws.cell(row=row, column=4, value="sec")
    shield_time_row = row
    row += 2

    ws.cell(row=row, column=2, value="HULL PHASE")
    ws.cell(row=row, column=2).font = SECTION_FONT
    row += 1

    ws.cell(row=row, column=2, value="Zone Hull HP:")
    ws.cell(row=row, column=3, value=f'=J{hull_hp_row}*VLOOKUP(C{target_zone_row},_Modifiers!N:R,2,FALSE)')
    apply_calc(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).number_format = '#,##0'
    zone_hull_row = row
    row += 1

    ws.cell(row=row, column=2, value="Zone Armor HP:")
    ws.cell(row=row, column=3, value=f'=J{armor_hp_row}*VLOOKUP(C{target_zone_row},_Modifiers!N:R,3,FALSE)')
    apply_armor(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).number_format = '#,##0'
    zone_armor_row = row
    row += 1

    ws.cell(row=row, column=2, value="Zone Thruster HP:")
    ws.cell(row=row, column=3, value=f'=J{thruster_total_row}*VLOOKUP(C{target_zone_row},_Modifiers!N:R,4,FALSE)')
    apply_thruster(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).number_format = '#,##0'
    zone_thruster_row = row
    row += 1

    ws.cell(row=row, column=2, value="Zone Component HP:")
    ws.cell(row=row, column=3, value=f'=(J{pp_hp_row}+J{cooler_hp_row}+J{shld_hp_row})*VLOOKUP(C{target_zone_row},_Modifiers!N:R,5,FALSE)')
    apply_component(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).number_format = '#,##0'
    zone_comp_row = row
    row += 1

    ws.cell(row=row, column=2, value="Total Zone HP:")
    ws.cell(row=row, column=3, value=f'=C{zone_hull_row}+C{zone_armor_row}+C{zone_thruster_row}+C{zone_comp_row}')
    apply_result(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).number_format = '#,##0'
    total_zone_hp_row = row
    row += 1

    ws.cell(row=row, column=2, value="Time to Destroy:")
    ws.cell(row=row, column=3, value=f'=IF(H{total_dps_row}>0,C{total_zone_hp_row}/H{total_dps_row},9999)')
    apply_result(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).number_format = '0.0'
    ws.cell(row=row, column=4, value="sec")
    hull_time_row = row
    row += 2

    # TOTAL TTK
    ws.cell(row=row, column=1, value="TOTAL TIME TO KILL")
    ws.cell(row=row, column=1).font = Font(bold=True, size=16, color="C00000")
    row += 2

    ws.cell(row=row, column=2, value="Shield:")
    ws.cell(row=row, column=3, value=f'=C{shield_time_row}')
    apply_result(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).number_format = '0.0'
    ws.cell(row=row, column=4, value="sec")
    row += 1

    ws.cell(row=row, column=2, value="Hull:")
    ws.cell(row=row, column=3, value=f'=C{hull_time_row}')
    apply_result(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).number_format = '0.0'
    ws.cell(row=row, column=4, value="sec")
    row += 2

    ws.cell(row=row, column=2, value="TOTAL TTK:")
    ws.cell(row=row, column=2).font = Font(bold=True, size=14)
    ttk_cell = ws.cell(row=row, column=3, value=f'=C{shield_time_row}+C{hull_time_row}')
    ttk_cell.fill = TTK_FILL
    ttk_cell.font = Font(bold=True, size=16, color="FFFFFF")
    ttk_cell.border = THIN_BORDER
    ttk_cell.alignment = Alignment(horizontal='center')
    ttk_cell.number_format = '0.0'
    ws.cell(row=row, column=4, value="seconds")
    ws.cell(row=row, column=4).font = Font(bold=True)

    # Column widths
    set_column_width(ws, 1, 3)
    set_column_width(ws, 2, 22)
    set_column_width(ws, 3, 22)
    set_column_width(ws, 4, 16)
    set_column_width(ws, 5, 10)
    set_column_width(ws, 6, 10)
    set_column_width(ws, 7, 8)
    set_column_width(ws, 8, 10)
    set_column_width(ws, 9, 16)
    set_column_width(ws, 10, 12)

    return ws


def create_instructions_sheet(wb):
    ws = wb.active
    ws.title = "Instructions"

    content = [
        ("STAR CITIZEN DAMAGE CALCULATOR v12", Font(bold=True, size=20, color="1F4E79")),
        ("Full Ship Parts System", Font(bold=True, size=14, color="C0392B")),
        ("", None),
        ("HOW TO USE:", Font(bold=True, size=12)),
        ("1. Select your ship from the dropdown", None),
        ("2. Choose your Combat Scenario", None),
        ("3. Select weapons for each slot", None),
        ("4. Select target ship and shield", None),
        ("5. Choose Target Zone for damage distribution", None),
        ("6. View TTK results", None),
    ]

    for row, (text, font) in enumerate(content, 1):
        cell = ws.cell(row=row, column=1, value=text)
        if font:
            cell.font = font

    set_column_width(ws, 1, 70)


def main():
    print("=" * 60)
    print("Creating Star Citizen Damage Calculator v12 (Fixed)")
    print("  - Filtered AI variants")
    print("  - Proper ship names")
    print("=" * 60)

    wb = openpyxl.Workbook()

    print("\nLoading data...")
    size_counts, weapon_total = create_weapons_data_sheet(wb)
    print(f"  Weapons: {weapon_total}")

    ship_names, ship_count = create_ships_data_sheet(wb)
    print(f"  Ships: {ship_count} (player-flyable only)")

    shield_names, shield_internal_names, shield_count, shield_size_counts = create_shields_data_sheet(wb)
    print(f"  Shields: {shield_count}")

    create_modifiers_sheet(wb)
    create_named_ranges(wb, size_counts, weapon_total, shield_size_counts, shield_count)

    print("\nCreating sheets...")
    create_instructions_sheet(wb)
    create_calculator_sheet(wb, ship_names, shield_names, shield_internal_names, ship_count, shield_count, size_counts, shield_size_counts)

    wb.move_sheet("Instructions", offset=0)
    wb.move_sheet("Calculator", offset=1)

    print(f"\nSaving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)

    print("\n" + "=" * 60)
    print("Damage Calculator v12 COMPLETE!")
    print("=" * 60)

    # Print sample ship names
    print("\nSample ship names:")
    for name in ship_names[:15]:
        print(f"  - {name}")


if __name__ == "__main__":
    main()
